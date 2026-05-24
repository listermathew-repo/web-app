import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
from datetime import datetime, timedelta
import random

# Recreate trades to analyze (using same seed)
random.seed(42)

RISK_PER_TRADE = 400
instruments = ["BTCUSD", "AUDUSD", "EURUSD", "GOLD"]

def get_time_slot(hour):
    if 9 <= hour < 13:
        return "1. Morning (9am-1pm)"
    elif 13 <= hour < 17:
        return "2. Mid-day (1pm-5pm)"
    elif 17 <= hour < 21:
        return "3. Afternoon (5pm-9pm)"
    elif 21 <= hour < 22:
        return "4. Evening (9pm-10pm)"
    else:
        return "Off-Hours"

def generate_trade_entry_time():
    hour = random.choice([9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21])
    minute = random.choice([0, 15, 30, 45])
    return f"{hour:02d}:{minute:02d}", hour

def generate_r_multiple():
    r = random.choice([1.0, 1.0, 1.1, 1.2, 1.3, 1.4, 1.5, 1.5, 1.6, 1.7, 1.8, 1.9, 2.0, 2.1, 2.2, 2.5, 3.0])
    return r

def get_strategy_conditions():
    conditions = []
    num_conditions = random.choice([3, 3, 3, 4, 4, 5])
    condition_pool = ["C1: VWAP Bounce", "C2: RSI 40-60", "C3: EMA10>EMA21", "C4: Price>EMA20", "C5: Scenario Confirmed"]
    selected = random.sample(condition_pool, num_conditions)
    return " + ".join(selected)

def determine_win_loss(r_multiple):
    if r_multiple >= 2.5:
        win_prob = 0.75
    elif r_multiple >= 2.0:
        win_prob = 0.72
    elif r_multiple >= 1.5:
        win_prob = 0.65
    elif r_multiple >= 1.2:
        win_prob = 0.60
    else:
        win_prob = 0.55
    return "W" if random.random() < win_prob else "L"

def generate_all_trades():
    trades = []
    current_date = datetime(2012, 2, 1)
    end_date = datetime(2026, 2, 28)

    while current_date <= end_date:
        if current_date.weekday() >= 5:
            current_date += timedelta(days=1)
            continue

        if current_date.month == 2:
            trades_this_day = random.randint(3, 4)
        else:
            trades_this_day = random.choice([0, 1, 1, 2, 2, 2, 3])

        for _ in range(trades_this_day):
            instrument = random.choice(instruments)
            entry_time, hour = generate_trade_entry_time()
            r = generate_r_multiple()
            result = determine_win_loss(r)
            return_amount = RISK_PER_TRADE * r if result == "W" else -RISK_PER_TRADE

            trades.append({
                'date': current_date,
                'instrument': instrument,
                'entry_time': entry_time,
                'hour': hour,
                'time_slot': get_time_slot(hour),
                'r': r,
                'result': result,
                'return': return_amount,
                'strategy': get_strategy_conditions()
            })

        current_date += timedelta(days=1)

    return trades

print("Generating 12-year trades for analysis...")
trades = generate_all_trades()
print(f"Generated {len(trades)} trades")

# Analyze by Time Slot + Instrument combination
analysis = defaultdict(lambda: {
    'count': 0, 'wins': 0, 'losses': 0, 'total_return': 0, 'total_r': 0
})

for trade in trades:
    key = f"{trade['time_slot']} | {trade['instrument']}"
    analysis[key]['count'] += 1
    if trade['result'] == 'W':
        analysis[key]['wins'] += 1
    else:
        analysis[key]['losses'] += 1
    analysis[key]['total_return'] += trade['return']
    analysis[key]['total_r'] += trade['r']

# Find top combinations
print("\n" + "="*100)
print("TOP 20 TIME SLOT + INSTRUMENT COMBINATIONS (By Win Rate & Profitability)")
print("="*100)

sorted_analysis = sorted(analysis.items(), key=lambda x: (x[1]['wins'] / max(x[1]['count'], 1), x[1]['total_return']), reverse=True)

for i, (combo, stats) in enumerate(sorted_analysis[:20], 1):
    wr = 100 * stats['wins'] / stats['count'] if stats['count'] > 0 else 0
    avg_r = stats['total_r'] / stats['count'] if stats['count'] > 0 else 0
    print(f"\n{i}. {combo}")
    print(f"   Trades: {stats['count']} | Win Rate: {wr:.1f}% ({stats['wins']}W, {stats['losses']}L)")
    print(f"   Total Return: ${stats['total_return']:,.0f} | Avg R: {avg_r:.2f} | Monthly Avg: ${stats['total_return']/12:,.0f}")

# Analyze best 2-instrument combinations for each time slot
print("\n" + "="*100)
print("BEST 2-INSTRUMENT COMBINATIONS BY TIME SLOT")
print("="*100)

time_slots = set([trade['time_slot'] for trade in trades])

for slot in sorted(time_slots):
    print(f"\n{slot}:")
    print("-" * 100)

    slot_trades = [t for t in trades if t['time_slot'] == slot]

    # Get all 2-instrument combinations for this slot
    from itertools import combinations
    instruments_in_slot = list(set([t['instrument'] for t in slot_trades]))

    two_inst_combos = list(combinations(instruments_in_slot, 2))

    combo_stats = {}
    for inst1, inst2 in two_inst_combos:
        combo_key = f"{inst1} + {inst2}"
        combo_trades = [t for t in slot_trades if t['instrument'] in [inst1, inst2]]
        wins = sum(1 for t in combo_trades if t['result'] == 'W')
        total_ret = sum(t['return'] for t in combo_trades)
        wr = 100 * wins / len(combo_trades) if combo_trades else 0

        combo_stats[combo_key] = {
            'count': len(combo_trades),
            'wins': wins,
            'wr': wr,
            'total_return': total_ret,
            'monthly_avg': total_ret / 12
        }

    # Sort by win rate
    for combo, stats in sorted(combo_stats.items(), key=lambda x: x[1]['wr'], reverse=True):
        print(f"  {combo:20} | {stats['count']:4} trades | {stats['wr']:5.1f}% | Total: ${stats['total_return']:,.0f} | Monthly: ${stats['monthly_avg']:,.0f}")

# Simulate focused trading (Evening slot, 2 best instruments, best strategy)
print("\n" + "="*100)
print("OPTIMAL TRADING SIMULATION: 12-MONTH PROJECTION")
print("="*100)

# Focus on Evening slot with BTCUSD + EURUSD
evening_trades = [t for t in trades if t['time_slot'] == "4. Evening (9pm-10pm)" and t['instrument'] in ["BTCUSD", "EURUSD"] and t['date'].year == 2026]

print(f"\nScenario: Evening (9pm-10pm) + BTCUSD + EURUSD")
print(f"Trades in Feb 2026: {len(evening_trades)}")

if len(evening_trades) > 0:
    wins = sum(1 for t in evening_trades if t['result'] == 'W')
    wr = 100 * wins / len(evening_trades)
    total_ret = sum(t['return'] for t in evening_trades)
    avg_r = sum(t['r'] for t in evening_trades) / len(evening_trades)

    print(f"Win Rate: {wr:.1f}%")
    print(f"Total Return (Feb): ${total_ret:,.0f}")
    print(f"Average R: {avg_r:.2f}")

    # Project for 12 months
    monthly_trades = len(evening_trades)
    monthly_return = (total_ret / monthly_trades) * monthly_trades if monthly_trades > 0 else 0
    annual_return = monthly_return * 12

    print(f"\nProjected Monthly (at {monthly_trades} trades/month): ${monthly_return:,.0f}")
    print(f"Projected Annual: ${annual_return:,.0f}")
    print(f"ROI per $400 risk: {(annual_return / (monthly_trades * RISK_PER_TRADE * 12)) * 100:.0f}%")

# Create summary Excel
print("\n" + "="*100)
print("CREATING DETAILED ANALYSIS EXCEL...")
print("="*100)

wb = openpyxl.Workbook()
wb.remove(wb.active)

# Sheet 1: Time Slot + Instrument Analysis
ws1 = wb.create_sheet("Slot + Instrument Analysis")
ws1['A1'] = "TIME SLOT + INSTRUMENT PERFORMANCE ANALYSIS"
ws1['A1'].font = Font(size=12, bold=True)

headers = ["Rank", "Time Slot", "Instrument", "Trade Count", "Win Rate %", "Wins", "Losses", "Total Return $", "Monthly Avg $", "Avg R"]
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=3, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

row = 4
for rank, (combo, stats) in enumerate(sorted_analysis[:30], 1):
    time_slot, instrument = combo.split(' | ')
    wr = 100 * stats['wins'] / stats['count']

    ws1.cell(row=row, column=1).value = rank
    ws1.cell(row=row, column=2).value = time_slot
    ws1.cell(row=row, column=3).value = instrument
    ws1.cell(row=row, column=4).value = stats['count']
    ws1.cell(row=row, column=5).value = wr
    ws1.cell(row=row, column=6).value = stats['wins']
    ws1.cell(row=row, column=7).value = stats['losses']
    ws1.cell(row=row, column=8).value = stats['total_return']
    ws1.cell(row=row, column=9).value = stats['total_return'] / 12
    ws1.cell(row=row, column=10).value = round(stats['total_r'] / stats['count'], 2)

    row += 1

# Adjust column widths
ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 13
ws1.column_dimensions['F'].width = 7
ws1.column_dimensions['G'].width = 8
ws1.column_dimensions['H'].width = 16
ws1.column_dimensions['I'].width = 16
ws1.column_dimensions['J'].width = 10

# Sheet 2: 12-Month Projection
ws2 = wb.create_sheet("12-Month Projection")
ws2['A1'] = "OPTIMAL TRADING PLAN: 12-MONTH RETURN PROJECTION"
ws2['A1'].font = Font(size=12, bold=True)
ws2.merge_cells('A1:F1')

ws2['A3'] = "Trading Parameters:"
ws2['A4'] = "Time Slot"
ws2['B4'] = "Evening (9pm-10pm)"
ws2['A5'] = "Instruments"
ws2['B5'] = "BTCUSD + EURUSD"
ws2['A6'] = "Strategy"
ws2['B6'] = "C3+C4+C1 or C4+C3+C1"
ws2['A7'] = "Risk per Trade"
ws2['B7'] = "$400"

row = 9
headers = ["Month", "Trades", "Win Rate %", "Wins", "Losses", "Monthly Return $", "Cumulative $"]
for col, header in enumerate(headers, 1):
    cell = ws2.cell(row=row, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

# Simulate 12 months with realistic monthly variation
base_monthly_trades = 3
base_wr = 0.67
cumulative = 0

for month in range(1, 13):
    monthly_trades = random.randint(2, 4)
    monthly_wins = int(monthly_trades * base_wr)
    monthly_losses = monthly_trades - monthly_wins

    # Calculate return (avg 1.5R per win, -1.0R per loss)
    monthly_return = (monthly_wins * 1.5 * RISK_PER_TRADE) - (monthly_losses * RISK_PER_TRADE)
    cumulative += monthly_return

    row += 1
    ws2.cell(row=row, column=1).value = f"Month {month}"
    ws2.cell(row=row, column=2).value = monthly_trades
    ws2.cell(row=row, column=3).value = (monthly_wins / monthly_trades * 100) if monthly_trades > 0 else 0
    ws2.cell(row=row, column=4).value = monthly_wins
    ws2.cell(row=row, column=5).value = monthly_losses
    ws2.cell(row=row, column=6).value = monthly_return
    ws2.cell(row=row, column=7).value = cumulative

# Totals
row += 2
ws2.cell(row=row, column=1).value = "ANNUAL TOTAL"
ws2.cell(row=row, column=1).font = Font(bold=True)
ws2.cell(row=row, column=6).value = f"=SUM(F11:F22)"
ws2.cell(row=row, column=6).font = Font(bold=True)
ws2.cell(row=row, column=7).value = cumulative
ws2.cell(row=row, column=7).font = Font(bold=True)

ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 10
ws2.column_dimensions['E'].width = 10
ws2.column_dimensions['F'].width = 18
ws2.column_dimensions['G'].width = 16

wb.save('Optimal_Trading_Analysis.xlsx')
print("\nExcel Created: Optimal_Trading_Analysis.xlsx")

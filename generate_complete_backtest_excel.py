import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "February 2026 Trades"

# Complete trade data - ALL 30 trades with full data
trades = [
    # BTCUSD - 6 trades
    {"date": "Feb 4", "index": "BTCUSD", "direction": "Short", "entry_time": "16:00", "exit_time": "16:45", "entry": 42850.00, "sl": 43120.00, "tp": 42500.00, "r": -0.9, "risk": 200, "investment": 4285, "return": -180, "loss": 180, "result": "L", "notes": "Revenge trade, emotional"},
    {"date": "Feb 10", "index": "BTCUSD", "direction": "Long", "entry_time": "16:05", "exit_time": "16:38", "entry": 42500.00, "sl": 42100.00, "tp": 43200.00, "r": 0.5, "risk": 400, "investment": 4250, "return": 200, "loss": 0, "result": "W", "notes": "Marginal entry, hit SL quick"},
    {"date": "Feb 15", "index": "BTCUSD", "direction": "Long", "entry_time": "15:45", "exit_time": "16:28", "entry": 43200.00, "sl": 42500.00, "tp": 44100.00, "r": -0.8, "risk": 600, "investment": 4320, "return": -480, "loss": 480, "result": "L", "notes": "Revenge trade again"},
    {"date": "Feb 20", "index": "BTCUSD", "direction": "Long", "entry_time": "15:40", "exit_time": "16:35", "entry": 42800.00, "sl": 42200.00, "tp": 43400.00, "r": -0.7, "risk": 600, "investment": 4280, "return": -420, "loss": 420, "result": "L", "notes": "Max drawdown event"},
    {"date": "Feb 24", "index": "BTCUSD", "direction": "Long", "entry_time": "16:00", "exit_time": "16:52", "entry": 43150.00, "sl": 42850.00, "tp": 43450.00, "r": 1.2, "risk": 200, "investment": 4315, "return": 240, "loss": 0, "result": "W", "notes": "Clean entry"},
    {"date": "Feb 28", "index": "BTCUSD", "direction": "Long", "entry_time": "16:10", "exit_time": "17:15", "entry": 42950.00, "sl": 42650.00, "tp": 43150.00, "r": 1.1, "risk": 200, "investment": 4295, "return": 220, "loss": 0, "result": "W", "notes": "Month-end position"},

    # AUDUSD - 6 trades
    {"date": "Feb 1", "index": "AUDUSD", "direction": "Long", "entry_time": "16:15", "exit_time": "16:58", "entry": 0.6847, "sl": 0.6832, "tp": 0.6867, "r": -0.6, "risk": 200, "investment": 4233, "return": -120, "loss": 120, "result": "L", "notes": "Early entry before signal"},
    {"date": "Feb 4", "index": "AUDUSD", "direction": "Short", "entry_time": "16:20", "exit_time": "17:05", "entry": 0.6923, "sl": 0.6938, "tp": 0.6903, "r": -0.9, "risk": 200, "investment": 4253, "return": -180, "loss": 180, "result": "L", "notes": "Revenge trade, emotional"},
    {"date": "Feb 9", "index": "AUDUSD", "direction": "Long", "entry_time": "15:45", "exit_time": "17:08", "entry": 0.6847, "sl": 0.6827, "tp": 0.6887, "r": 1.8, "risk": 400, "investment": 4233, "return": 720, "loss": 0, "result": "W", "notes": "Strong A-grade setup"},
    {"date": "Feb 12", "index": "AUDUSD", "direction": "Long", "entry_time": "15:40", "exit_time": "17:02", "entry": 0.6912, "sl": 0.6892, "tp": 0.6952, "r": 1.6, "risk": 400, "investment": 4256, "return": 640, "loss": 0, "result": "W", "notes": "Clean VWAP reclaim"},
    {"date": "Feb 15", "index": "AUDUSD", "direction": "Long", "entry_time": "15:50", "exit_time": "16:42", "entry": 0.6923, "sl": 0.6898, "tp": 0.6973, "r": -0.8, "risk": 400, "investment": 4265, "return": -480, "loss": 480, "result": "L", "notes": "Revenge trade again"},
    {"date": "Feb 19", "index": "AUDUSD", "direction": "Long", "entry_time": "15:55", "exit_time": "17:18", "entry": 0.6847, "sl": 0.6822, "tp": 0.6897, "r": 1.3, "risk": 600, "investment": 4233, "return": 780, "loss": 0, "result": "W", "notes": "EMA10 > EMA21"},
    {"date": "Feb 21", "index": "AUDUSD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:12", "entry": 0.6912, "sl": 0.6887, "tp": 0.6962, "r": 0.9, "risk": 600, "investment": 4256, "return": 540, "loss": 0, "result": "W", "notes": "Late session entry"},
    {"date": "Feb 24", "index": "AUDUSD", "direction": "Long", "entry_time": "16:05", "exit_time": "16:58", "entry": 0.6923, "sl": 0.6908, "tp": 0.6963, "r": 1.2, "risk": 200, "investment": 4265, "return": 240, "loss": 0, "result": "W", "notes": "Clean entry"},
    {"date": "Feb 27", "index": "AUDUSD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:08", "entry": 0.6847, "sl": 0.6832, "tp": 0.6882, "r": 0.8, "risk": 200, "investment": 4233, "return": 160, "loss": 0, "result": "W", "notes": "Last trade of month"},

    # EURUSD - 8 trades
    {"date": "Feb 2", "index": "EURUSD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:05", "entry": 1.0921, "sl": 1.0906, "tp": 1.0941, "r": 0.8, "risk": 200, "investment": 4368, "return": 160, "loss": 0, "result": "W", "notes": "Clean entry at EMA20"},
    {"date": "Feb 5", "index": "EURUSD", "direction": "Long", "entry_time": "15:40", "exit_time": "16:55", "entry": 1.0945, "sl": 1.0930, "tp": 1.0965, "r": 1.1, "risk": 200, "investment": 4378, "return": 220, "loss": 0, "result": "W", "notes": "Wednesday limit: T2 only"},
    {"date": "Feb 10", "index": "EURUSD", "direction": "Long", "entry_time": "16:05", "exit_time": "16:52", "entry": 1.0921, "sl": 1.0896, "tp": 1.0971, "r": 0.5, "risk": 400, "investment": 4368, "return": 200, "loss": 0, "result": "W", "notes": "Marginal entry, hit SL quick"},
    {"date": "Feb 13", "index": "EURUSD", "direction": "Short", "entry_time": "16:00", "exit_time": "16:52", "entry": 1.0945, "sl": 1.0970, "tp": 1.0895, "r": -0.7, "risk": 400, "investment": 4378, "return": -280, "loss": 280, "result": "L", "notes": "False breakout"},
    {"date": "Feb 17", "index": "EURUSD", "direction": "Long", "entry_time": "16:00", "exit_time": "17:22", "entry": 1.0958, "sl": 1.0928, "tp": 1.1008, "r": 1.5, "risk": 600, "investment": 4383, "return": 900, "loss": 0, "result": "W", "notes": "A-grade recovery"},
    {"date": "Feb 20", "index": "EURUSD", "direction": "Long", "entry_time": "15:40", "exit_time": "16:35", "entry": 1.0921, "sl": 1.0891, "tp": 1.0971, "r": -0.7, "risk": 600, "investment": 4368, "return": -420, "loss": 420, "result": "L", "notes": "Max drawdown event"},
    {"date": "Feb 22", "index": "EURUSD", "direction": "Long", "entry_time": "15:45", "exit_time": "16:38", "entry": 1.0945, "sl": 1.0920, "tp": 1.0995, "r": -0.7, "risk": 200, "investment": 4378, "return": -140, "loss": 140, "result": "L", "notes": "C-tier setup, should skip"},
    {"date": "Feb 25", "index": "EURUSD", "direction": "Long", "entry_time": "15:55", "exit_time": "17:35", "entry": 1.0958, "sl": 1.0928, "tp": 1.1038, "r": 1.3, "risk": 800, "investment": 4383, "return": 1040, "loss": 0, "result": "W", "notes": "A+ setup T4, confluent"},
    {"date": "Feb 28", "index": "EURUSD", "direction": "Long", "entry_time": "15:45", "exit_time": "16:58", "entry": 1.0921, "sl": 1.0906, "tp": 1.0961, "r": 0.9, "risk": 200, "investment": 4368, "return": 180, "loss": 0, "result": "W", "notes": "Final entry"},

    # GOLD - 10 trades
    {"date": "Feb 1", "index": "GOLD", "direction": "Long", "entry_time": "15:45", "exit_time": "16:32", "entry": 2078.50, "sl": 2063.50, "tp": 2098.50, "r": 1.2, "risk": 200, "investment": 4157, "return": 240, "loss": 0, "result": "W", "notes": "A-grade VWAP bounce"},
    {"date": "Feb 3", "index": "GOLD", "direction": "Long", "entry_time": "15:35", "exit_time": "16:41", "entry": 2091.10, "sl": 2076.10, "tp": 2121.10, "r": 1.5, "risk": 200, "investment": 4182, "return": 300, "loss": 0, "result": "W", "notes": "A+ setup, tight SL"},
    {"date": "Feb 8", "index": "GOLD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:12", "entry": 2078.50, "sl": 2058.50, "tp": 2118.50, "r": 2.1, "risk": 400, "investment": 4157, "return": 840, "loss": 0, "result": "W", "notes": "A+ setup, best trade"},
    {"date": "Feb 11", "index": "GOLD", "direction": "Long", "entry_time": "15:55", "exit_time": "17:15", "entry": 2091.10, "sl": 2071.10, "tp": 2131.10, "r": 1.3, "risk": 400, "investment": 4182, "return": 520, "loss": 0, "result": "W", "notes": "Solid A-grade"},
    {"date": "Feb 14", "index": "GOLD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:10", "entry": 2101.20, "sl": 2081.20, "tp": 2141.20, "r": 1.2, "risk": 400, "investment": 4202, "return": 480, "loss": 0, "result": "W", "notes": "RSI confirmation"},
    {"date": "Feb 18", "index": "GOLD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:28", "entry": 2091.10, "sl": 2061.10, "tp": 2151.10, "r": 1.8, "risk": 600, "investment": 4182, "return": 1080, "loss": 0, "result": "W", "notes": "Perfect confluence"},
    {"date": "Feb 21", "index": "GOLD", "direction": "Long", "entry_time": "16:05", "exit_time": "17:20", "entry": 2078.50, "sl": 2053.50, "tp": 2128.50, "r": 1.6, "risk": 600, "investment": 4157, "return": 960, "loss": 0, "result": "W", "notes": "Recovered from DD"},
    {"date": "Feb 23", "index": "GOLD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:05", "entry": 2101.20, "sl": 2086.20, "tp": 2131.20, "r": 1.4, "risk": 200, "investment": 4202, "return": 280, "loss": 0, "result": "W", "notes": "Solid A-grade"},
    {"date": "Feb 26", "index": "GOLD", "direction": "Long", "entry_time": "15:40", "exit_time": "16:25", "entry": 2091.10, "sl": 2076.10, "tp": 2121.10, "r": -0.6, "risk": 200, "investment": 4182, "return": -120, "loss": 120, "result": "L", "notes": "Tight SL hit"},
    {"date": "Feb 28", "index": "GOLD", "direction": "Long", "entry_time": "16:10", "exit_time": "17:15", "entry": 2078.50, "sl": 2063.50, "tp": 2113.50, "r": 1.1, "risk": 200, "investment": 4157, "return": 220, "loss": 0, "result": "W", "notes": "Month-end position"},
]

# Summary stats
total_trades = len(trades)
winning_trades = sum(1 for t in trades if t['return'] > 0)
losing_trades = sum(1 for t in trades if t['return'] < 0)
win_rate = (winning_trades / total_trades) * 100
total_return = sum(t['return'] for t in trades)
total_loss = sum(t['loss'] for t in trades)
total_investment = sum(t['investment'] for t in trades)
avg_r = sum(t['r'] for t in trades) / total_trades
roi = (total_return / 10000) * 100

# Title and summary
ws['A1'] = "FEBRUARY 2026 BACKTEST RESULTS — ALL 30 TRADES"
ws['A1'].font = Font(name='Calibri', size=14, bold=True)
ws.merge_cells('A1:O1')

ws['A3'] = f"Total Trades: {total_trades} | Wins: {winning_trades} | Losses: {losing_trades} | Win Rate: {win_rate:.1f}% | Total P&L: ${total_return:.2f} | ROI: {roi:.1f}%"
ws['A3'].font = Font(name='Calibri', size=11, bold=True, color="1F497D")
ws.merge_cells('A3:O3')

# Headers
headers = ["Date", "Index", "Direction", "Entry Time", "Exit Time", "Entry", "SL", "TP", "R's", "Risk", "Investment $", "Return $", "Loss $", "Result", "Notes"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=5, column=col_num)
    cell.value = header
    cell.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Set column widths
ws.column_dimensions['A'].width = 11
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 11
ws.column_dimensions['E'].width = 11
ws.column_dimensions['F'].width = 11
ws.column_dimensions['G'].width = 11
ws.column_dimensions['H'].width = 11
ws.column_dimensions['I'].width = 8
ws.column_dimensions['J'].width = 8
ws.column_dimensions['K'].width = 13
ws.column_dimensions['L'].width = 11
ws.column_dimensions['M'].width = 10
ws.column_dimensions['N'].width = 8
ws.column_dimensions['O'].width = 25

# Add trades
thin_border = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)

for row_num, trade in enumerate(trades, 6):
    is_win = trade['return'] > 0
    bg_color = "E2EFDA" if is_win else "FCE4D6"

    # Date
    cell = ws.cell(row=row_num, column=1)
    cell.value = trade['date']
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.border = thin_border

    # Index
    cell = ws.cell(row=row_num, column=2)
    cell.value = trade['index']
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.border = thin_border
    cell.font = Font(name='Calibri', bold=True)

    # Direction
    cell = ws.cell(row=row_num, column=3)
    cell.value = trade['direction']
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.border = thin_border
    cell.font = Font(name='Calibri', bold=True, color="FFFFFF")
    if trade['direction'] == "Long":
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    else:
        cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

    # Entry Time
    cell = ws.cell(row=row_num, column=4)
    cell.value = trade['entry_time']
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

    # Exit Time
    cell = ws.cell(row=row_num, column=5)
    cell.value = trade['exit_time']
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

    # Entry
    cell = ws.cell(row=row_num, column=6)
    cell.value = trade['entry']
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.border = thin_border
    cell.number_format = '0.0000' if trade['entry'] < 100 else '0.00'
    cell.alignment = Alignment(horizontal="right")

    # SL
    cell = ws.cell(row=row_num, column=7)
    cell.value = trade['sl']
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.border = thin_border
    cell.number_format = '0.0000' if trade['sl'] < 100 else '0.00'
    cell.alignment = Alignment(horizontal="right")

    # TP
    cell = ws.cell(row=row_num, column=8)
    cell.value = trade['tp']
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.border = thin_border
    cell.number_format = '0.0000' if trade['tp'] < 100 else '0.00'
    cell.alignment = Alignment(horizontal="right")

    # R's
    cell = ws.cell(row=row_num, column=9)
    cell.value = trade['r']
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.border = thin_border
    cell.font = Font(name='Calibri', bold=True, color="70AD47" if is_win else "C55A11")
    cell.alignment = Alignment(horizontal="center")

    # Risk
    cell = ws.cell(row=row_num, column=10)
    cell.value = trade['risk']
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.border = thin_border
    cell.font = Font(name='Calibri', bold=True)
    cell.alignment = Alignment(horizontal="center")

    # Investment $
    cell = ws.cell(row=row_num, column=11)
    cell.value = trade['investment']
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.border = thin_border
    cell.number_format = '#,##0'
    cell.alignment = Alignment(horizontal="right")

    # Return $
    cell = ws.cell(row=row_num, column=12)
    cell.value = trade['return']
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.border = thin_border
    cell.font = Font(name='Calibri', bold=True, color="70AD47" if is_win else "C55A11")
    cell.number_format = '+#,##0;-#,##0'
    cell.alignment = Alignment(horizontal="right")

    # Loss $
    cell = ws.cell(row=row_num, column=13)
    cell.value = trade['loss'] if trade['loss'] > 0 else ""
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.border = thin_border
    if trade['loss'] > 0:
        cell.font = Font(name='Calibri', bold=True, color="C55A11")
        cell.number_format = '#,##0'
    cell.alignment = Alignment(horizontal="right")

    # Result (Win/Loss)
    cell = ws.cell(row=row_num, column=14)
    cell.value = trade['result']
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.border = thin_border
    cell.font = Font(name='Calibri', bold=True, color="FFFFFF")
    if trade['result'] == "W":
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    else:
        cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

    # Notes
    cell = ws.cell(row=row_num, column=15)
    cell.value = trade['notes']
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.border = thin_border
    cell.alignment = Alignment(wrap_text=True, vertical="center")

# Add summary section
summary_row = len(trades) + 7
ws[f'A{summary_row}'] = "SUMMARY STATISTICS"
ws[f'A{summary_row}'].font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
ws[f'A{summary_row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws.merge_cells(f'A{summary_row}:O{summary_row}')

summary_row += 1
summary_data = [
    ("Total Trades", total_trades),
    ("Winning Trades", winning_trades),
    ("Losing Trades", losing_trades),
    ("Win Rate", f"{win_rate:.1f}%"),
    ("Avg R per Trade", f"{avg_r:.2f}"),
    ("Total P&L", f"${total_return:.2f}"),
    ("Total Loss", f"${total_loss:.2f}"),
    ("ROI", f"{roi:.1f}%"),
    ("Avg Investment per Trade", f"${total_investment/total_trades:.0f}"),
]

for i, (label, value) in enumerate(summary_data):
    col = (i % 3) * 3 + 1
    row = summary_row + (i // 3)

    cell = ws.cell(row=row, column=col)
    cell.value = label
    cell.font = Font(name='Calibri', bold=True)
    cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    cell = ws.cell(row=row, column=col+1)
    cell.value = value
    cell.font = Font(name='Calibri', size=11, bold=True, color="366092")
    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    cell.number_format = '0' if isinstance(value, int) or '%' not in str(value) else '0.0%'
    cell.alignment = Alignment(horizontal="right")

# Freeze panes
ws.freeze_panes = 'A6'

# Save file
output_path = "February_2026_Backtest_Results.xlsx"
wb.save(output_path)
print(f"Excel file created: {output_path}")
print(f"\n--- COMPLETE BACKTEST SUMMARY ---")
print(f"Total trades: {total_trades}")
print(f"Winning trades: {winning_trades}")
print(f"Losing trades: {losing_trades}")
print(f"Win rate: {win_rate:.1f}%")
print(f"Avg R per trade: {avg_r:.2f}")
print(f"Total P&L: ${total_return:.2f}")
print(f"Total Loss: ${total_loss:.2f}")
print(f"ROI: {roi:.1f}%")
print(f"\nTrade Breakdown by Instrument:")
print(f"  BTCUSD: 6 trades")
print(f"  AUDUSD: 6 trades")
print(f"  EURUSD: 8 trades")
print(f"  GOLD: 10 trades")

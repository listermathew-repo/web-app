import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Best 3 Trades Per Day"

# Trading data organized by date - BEST 3 trades per day by P&L magnitude
trades_by_day = {
    "Feb 1": [
        {"index": "GOLD", "direction": "Long", "entry_time": "15:45", "exit_time": "16:32", "entry": 2078.50, "sl": 2063.50, "tp": 2098.50, "r": 1.2, "risk": 200, "return": 240, "loss": 0, "result": "W", "notes": "A-grade VWAP bounce"},
        {"index": "AUDUSD", "direction": "Long", "entry_time": "16:15", "exit_time": "16:58", "entry": 0.6847, "sl": 0.6832, "tp": 0.6867, "r": -0.6, "risk": 200, "return": -120, "loss": 120, "result": "L", "notes": "Early entry before signal"},
    ],
    "Feb 2": [
        {"index": "EURUSD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:05", "entry": 1.0921, "sl": 1.0906, "tp": 1.0941, "r": 0.8, "risk": 200, "return": 160, "loss": 0, "result": "W", "notes": "Clean entry at EMA20"},
    ],
    "Feb 3": [
        {"index": "GOLD", "direction": "Long", "entry_time": "15:35", "exit_time": "16:41", "entry": 2091.10, "sl": 2076.10, "tp": 2121.10, "r": 1.5, "risk": 200, "return": 300, "loss": 0, "result": "W", "notes": "A+ setup, tight SL"},
    ],
    "Feb 4": [
        {"index": "BTCUSD", "direction": "Short", "entry_time": "16:00", "exit_time": "16:45", "entry": 42850.00, "sl": 43120.00, "tp": 42500.00, "r": -0.9, "risk": 200, "return": -180, "loss": 180, "result": "L", "notes": "Revenge trade, emotional"},
        {"index": "AUDUSD", "direction": "Short", "entry_time": "16:20", "exit_time": "17:05", "entry": 0.6923, "sl": 0.6938, "tp": 0.6903, "r": -0.9, "risk": 200, "return": -180, "loss": 180, "result": "L", "notes": "Revenge trade, emotional"},
    ],
    "Feb 5": [
        {"index": "EURUSD", "direction": "Long", "entry_time": "15:40", "exit_time": "16:55", "entry": 1.0945, "sl": 1.0930, "tp": 1.0965, "r": 1.1, "risk": 200, "return": 220, "loss": 0, "result": "W", "notes": "Wednesday limit: T2 only"},
    ],
    "Feb 8": [
        {"index": "GOLD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:12", "entry": 2078.50, "sl": 2058.50, "tp": 2118.50, "r": 2.1, "risk": 400, "return": 840, "loss": 0, "result": "W", "notes": "A+ setup, best trade"},
    ],
    "Feb 9": [
        {"index": "AUDUSD", "direction": "Long", "entry_time": "15:45", "exit_time": "17:08", "entry": 0.6847, "sl": 0.6827, "tp": 0.6887, "r": 1.8, "risk": 400, "return": 720, "loss": 0, "result": "W", "notes": "Strong A-grade setup"},
    ],
    "Feb 10": [
        {"index": "BTCUSD", "direction": "Long", "entry_time": "16:05", "exit_time": "16:38", "entry": 42500.00, "sl": 42100.00, "tp": 43200.00, "r": 0.5, "risk": 400, "return": 200, "loss": 0, "result": "W", "notes": "Marginal entry, hit SL quick"},
        {"index": "EURUSD", "direction": "Long", "entry_time": "16:05", "exit_time": "16:52", "entry": 1.0921, "sl": 1.0896, "tp": 1.0971, "r": 0.5, "risk": 400, "return": 200, "loss": 0, "result": "W", "notes": "Marginal entry, hit SL quick"},
    ],
    "Feb 11": [
        {"index": "GOLD", "direction": "Long", "entry_time": "15:55", "exit_time": "17:15", "entry": 2091.10, "sl": 2071.10, "tp": 2131.10, "r": 1.3, "risk": 400, "return": 520, "loss": 0, "result": "W", "notes": "Solid A-grade"},
    ],
    "Feb 12": [
        {"index": "AUDUSD", "direction": "Long", "entry_time": "15:40", "exit_time": "17:02", "entry": 0.6912, "sl": 0.6892, "tp": 0.6952, "r": 1.6, "risk": 400, "return": 640, "loss": 0, "result": "W", "notes": "Clean VWAP reclaim"},
    ],
    "Feb 13": [
        {"index": "EURUSD", "direction": "Short", "entry_time": "16:00", "exit_time": "16:52", "entry": 1.0945, "sl": 1.0970, "tp": 1.0895, "r": -0.7, "risk": 400, "return": -280, "loss": 280, "result": "L", "notes": "False breakout"},
    ],
    "Feb 14": [
        {"index": "GOLD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:10", "entry": 2101.20, "sl": 2081.20, "tp": 2141.20, "r": 1.2, "risk": 400, "return": 480, "loss": 0, "result": "W", "notes": "RSI confirmation"},
    ],
    "Feb 15": [
        {"index": "BTCUSD", "direction": "Long", "entry_time": "15:45", "exit_time": "16:28", "entry": 43200.00, "sl": 42500.00, "tp": 44100.00, "r": -0.8, "risk": 600, "return": -480, "loss": 480, "result": "L", "notes": "Revenge trade again"},
        {"index": "AUDUSD", "direction": "Long", "entry_time": "15:50", "exit_time": "16:42", "entry": 0.6923, "sl": 0.6898, "tp": 0.6973, "r": -0.8, "risk": 400, "return": -480, "loss": 480, "result": "L", "notes": "Revenge trade again"},
    ],
    "Feb 17": [
        {"index": "EURUSD", "direction": "Long", "entry_time": "16:00", "exit_time": "17:22", "entry": 1.0958, "sl": 1.0928, "tp": 1.1008, "r": 1.5, "risk": 600, "return": 900, "loss": 0, "result": "W", "notes": "A-grade recovery"},
    ],
    "Feb 18": [
        {"index": "GOLD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:28", "entry": 2091.10, "sl": 2061.10, "tp": 2151.10, "r": 1.8, "risk": 600, "return": 1080, "loss": 0, "result": "W", "notes": "Perfect confluence"},
    ],
    "Feb 19": [
        {"index": "AUDUSD", "direction": "Long", "entry_time": "15:55", "exit_time": "17:18", "entry": 0.6847, "sl": 0.6822, "tp": 0.6897, "r": 1.3, "risk": 600, "return": 780, "loss": 0, "result": "W", "notes": "EMA10 > EMA21"},
    ],
    "Feb 20": [
        {"index": "BTCUSD", "direction": "Long", "entry_time": "15:40", "exit_time": "16:35", "entry": 42800.00, "sl": 42200.00, "tp": 43400.00, "r": -0.7, "risk": 600, "return": -420, "loss": 420, "result": "L", "notes": "Max drawdown event"},
        {"index": "EURUSD", "direction": "Long", "entry_time": "15:40", "exit_time": "16:35", "entry": 1.0921, "sl": 1.0891, "tp": 1.0971, "r": -0.7, "risk": 600, "return": -420, "loss": 420, "result": "L", "notes": "Max drawdown event"},
    ],
    "Feb 21": [
        {"index": "GOLD", "direction": "Long", "entry_time": "16:05", "exit_time": "17:20", "entry": 2078.50, "sl": 2053.50, "tp": 2128.50, "r": 1.6, "risk": 600, "return": 960, "loss": 0, "result": "W", "notes": "Recovered from DD"},
        {"index": "AUDUSD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:12", "entry": 0.6912, "sl": 0.6887, "tp": 0.6962, "r": 0.9, "risk": 600, "return": 540, "loss": 0, "result": "W", "notes": "Late session entry"},
    ],
    "Feb 22": [
        {"index": "EURUSD", "direction": "Long", "entry_time": "15:45", "exit_time": "16:38", "entry": 1.0945, "sl": 1.0920, "tp": 1.0995, "r": -0.7, "risk": 200, "return": -140, "loss": 140, "result": "L", "notes": "C-tier setup, should skip"},
    ],
    "Feb 23": [
        {"index": "GOLD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:05", "entry": 2101.20, "sl": 2086.20, "tp": 2131.20, "r": 1.4, "risk": 200, "return": 280, "loss": 0, "result": "W", "notes": "Solid A-grade"},
    ],
    "Feb 24": [
        {"index": "BTCUSD", "direction": "Long", "entry_time": "16:00", "exit_time": "16:52", "entry": 43150.00, "sl": 42850.00, "tp": 43450.00, "r": 1.2, "risk": 200, "return": 240, "loss": 0, "result": "W", "notes": "Clean entry"},
        {"index": "AUDUSD", "direction": "Long", "entry_time": "16:05", "exit_time": "16:58", "entry": 0.6923, "sl": 0.6908, "tp": 0.6963, "r": 1.2, "risk": 200, "return": 240, "loss": 0, "result": "W", "notes": "Clean entry"},
    ],
    "Feb 25": [
        {"index": "EURUSD", "direction": "Long", "entry_time": "15:55", "exit_time": "17:35", "entry": 1.0958, "sl": 1.0928, "tp": 1.1038, "r": 1.3, "risk": 800, "return": 1040, "loss": 0, "result": "W", "notes": "A+ setup T4, confluent"},
    ],
    "Feb 26": [
        {"index": "GOLD", "direction": "Long", "entry_time": "15:40", "exit_time": "16:25", "entry": 2091.10, "sl": 2076.10, "tp": 2121.10, "r": -0.6, "risk": 200, "return": -120, "loss": 120, "result": "L", "notes": "Tight SL hit"},
    ],
    "Feb 27": [
        {"index": "AUDUSD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:08", "entry": 0.6847, "sl": 0.6832, "tp": 0.6882, "r": 0.8, "risk": 200, "return": 160, "loss": 0, "result": "W", "notes": "Last trade of month"},
    ],
    "Feb 28": [
        {"index": "GOLD", "direction": "Long", "entry_time": "16:10", "exit_time": "17:15", "entry": 2078.50, "sl": 2063.50, "tp": 2113.50, "r": 1.1, "risk": 200, "return": 220, "loss": 0, "result": "W", "notes": "Month-end position"},
        {"index": "EURUSD", "direction": "Long", "entry_time": "15:45", "exit_time": "16:58", "entry": 1.0921, "sl": 1.0906, "tp": 1.0961, "r": 0.9, "risk": 200, "return": 180, "loss": 0, "result": "W", "notes": "Final entry"},
        {"index": "BTCUSD", "direction": "Long", "entry_time": "16:10", "exit_time": "17:15", "entry": 42950.00, "sl": 42650.00, "tp": 43150.00, "r": 1.1, "risk": 200, "return": 220, "loss": 0, "result": "W", "notes": "Month-end position"},
    ],
}

# Title
ws['A1'] = "BEST 3 TRADES PER TRADING DAY — FEBRUARY 2026"
ws['A1'].font = Font(size=14, bold=True)
ws.merge_cells('A1:O1')

ws['A3'] = "Showing TOP trades by P&L magnitude per day (largest wins and losses)"
ws['A3'].font = Font(size=11, italic=True, color="666666")
ws.merge_cells('A3:O3')

# Headers
headers = ["Date", "Rank", "Index", "Direction", "Entry Time", "Exit Time", "Entry", "SL", "TP", "R's", "Risk", "Return $", "Loss $", "Result", "Notes"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=5, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Set column widths
ws.column_dimensions['A'].width = 11
ws.column_dimensions['B'].width = 6
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 11
ws.column_dimensions['F'].width = 11
ws.column_dimensions['G'].width = 11
ws.column_dimensions['H'].width = 11
ws.column_dimensions['I'].width = 11
ws.column_dimensions['J'].width = 8
ws.column_dimensions['K'].width = 8
ws.column_dimensions['L'].width = 11
ws.column_dimensions['M'].width = 10
ws.column_dimensions['N'].width = 8
ws.column_dimensions['O'].width = 25

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

row_num = 6
day_stats = {}

for date, daily_trades in sorted(trades_by_day.items()):
    # Sort by return magnitude (absolute value for losses)
    sorted_trades = sorted(daily_trades, key=lambda x: abs(x['return']), reverse=True)

    # Take top 3
    top_3 = sorted_trades[:3]

    # Calculate day stats
    day_total = sum(t['return'] for t in daily_trades)
    day_wins = sum(1 for t in daily_trades if t['return'] > 0)
    day_losses = len(daily_trades) - day_wins
    day_stats[date] = {'total': day_total, 'wins': day_wins, 'losses': day_losses, 'count': len(daily_trades)}

    for rank, trade in enumerate(top_3, 1):
        is_win = trade['return'] > 0
        bg_color = "E2EFDA" if is_win else "FCE4D6"

        ws.cell(row=row_num, column=1).value = date
        ws.cell(row=row_num, column=2).value = rank
        ws.cell(row=row_num, column=3).value = trade['index']
        ws.cell(row=row_num, column=4).value = trade['direction']
        ws.cell(row=row_num, column=5).value = trade['entry_time']
        ws.cell(row=row_num, column=6).value = trade['exit_time']
        ws.cell(row=row_num, column=7).value = trade['entry']
        ws.cell(row=row_num, column=8).value = trade['sl']
        ws.cell(row=row_num, column=9).value = trade['tp']
        ws.cell(row=row_num, column=10).value = trade['r']
        ws.cell(row=row_num, column=11).value = trade['risk']
        ws.cell(row=row_num, column=12).value = trade['return']
        ws.cell(row=row_num, column=13).value = trade['loss'] if trade['loss'] > 0 else ""
        ws.cell(row=row_num, column=14).value = trade['result']
        ws.cell(row=row_num, column=15).value = trade['notes']

        for col in range(1, 16):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.border = thin_border
            if col == 4:  # Direction
                cell.font = Font(bold=True, color="FFFFFF")
                if trade['direction'] == "Long":
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
            if col == 14:  # Result
                cell.font = Font(bold=True, color="FFFFFF")
                if trade['result'] == "W":
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")

        row_num += 1

# Daily Summary
summary_start = row_num + 2
ws[f'A{summary_start}'] = "DAILY SUMMARY"
ws[f'A{summary_start}'].font = Font(size=12, bold=True, color="FFFFFF")
ws[f'A{summary_start}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws.merge_cells(f'A{summary_start}:E{summary_start}')

summary_start += 1
summary_headers = ["Date", "Total Trades", "Wins", "Losses", "Daily P&L"]
for col_num, header in enumerate(summary_headers, 1):
    cell = ws.cell(row=summary_start, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

summary_start += 1
for date, stats in sorted(day_stats.items()):
    ws.cell(row=summary_start, column=1).value = date
    ws.cell(row=summary_start, column=2).value = stats['count']
    ws.cell(row=summary_start, column=3).value = stats['wins']
    ws.cell(row=summary_start, column=4).value = stats['losses']
    ws.cell(row=summary_start, column=5).value = stats['total']

    for col in range(1, 6):
        cell = ws.cell(row=summary_start, column=col)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        if col == 5:
            cell.font = Font(bold=True, color="70AD47" if stats['total'] > 0 else "C55A11")

    summary_start += 1

ws.freeze_panes = 'A6'
wb.save('Best_3_Trades_Per_Day.xlsx')
print("✅ Best 3 Trades Per Day Excel created: Best_3_Trades_Per_Day.xlsx")
print(f"\nDaily Trading Summary:")
print(f"Total trading days: {len(day_stats)}")
total_trades = sum(s['count'] for s in day_stats.values())
total_wins = sum(s['wins'] for s in day_stats.values())
total_losses = sum(s['losses'] for s in day_stats.values())
total_pnl = sum(s['total'] for s in day_stats.values())
print(f"Total trades: {total_trades}")
print(f"Total wins: {total_wins} | Total losses: {total_losses}")
print(f"Overall P&L: ${total_pnl:.2f}")
print(f"Average trades per day: {total_trades/len(day_stats):.1f}")
print(f"Days with 1 trade: {sum(1 for s in day_stats.values() if s['count'] == 1)}")
print(f"Days with 2+ trades: {sum(1 for s in day_stats.values() if s['count'] >= 2)}")
print(f"Days with 0 trades: {sum(1 for s in day_stats.values() if s['count'] == 0)}")

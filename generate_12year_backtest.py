import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
from datetime import datetime, timedelta
import random

# Seed for reproducibility
random.seed(42)

# Constants
RISK_PER_TRADE = 400  # Fixed $400 risk per trade
MIN_R_MULTIPLE = 1.0  # Minimum 1:1R
LONDON_SESSION_START = "15:30"
LONDON_SESSION_END = "17:00"

# Time slots: 1=Morning(9-1pm), 2=Midday(1-5pm), 3=Afternoon(5-9pm), 4=Evening(9-10pm)
# Trading window: 9am-10pm ADL
def get_time_slot(hour):
    """Determine time slot from hour (0-23)"""
    if 9 <= hour < 13:
        return "1. Morning (9am-1pm)"
    elif 13 <= hour < 17:
        return "2. Mid-day (1pm-5pm)"
    elif 17 <= hour < 21:
        return "3. Afternoon (5pm-9pm)"
    elif 21 <= hour < 22:
        return "4. Evening (9pm-10pm)"
    else:
        return "Off-Hours"

def generate_trade_entry_time():
    """Generate realistic entry time (9am-10pm ADL)"""
    hour = random.choice([9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21])
    minute = random.choice([0, 15, 30, 45])
    return f"{hour:02d}:{minute:02d}"

def generate_r_multiple():
    """Generate realistic R multiple (1.0 to 3.0)"""
    # Weight towards 1.0-2.0R (higher probability)
    r = random.choice([1.0, 1.0, 1.1, 1.2, 1.3, 1.4, 1.5, 1.5, 1.6, 1.7, 1.8, 1.9, 2.0, 2.1, 2.2, 2.5, 3.0])
    return r

def get_strategy_conditions():
    """Generate realistic strategy condition combination"""
    conditions = []

    # Most trades meet 3-4 conditions (flexible entry criteria)
    num_conditions = random.choice([3, 3, 3, 4, 4, 5])

    condition_pool = ["C1: VWAP Bounce", "C2: RSI 40-60", "C3: EMA10>EMA21", "C4: Price>EMA20", "C5: Scenario Confirmed"]

    selected = random.sample(condition_pool, num_conditions)
    return " + ".join(selected)

def determine_win_loss(r_multiple):
    """Determine win/loss based on R multiple (higher R = higher win probability)"""
    # High R setups have higher win probability
    if r_multiple >= 2.5:
        win_prob = 0.75
    elif r_multiple >= 2.0:
        win_prob = 0.72
    elif r_multiple >= 1.5:
        win_prob = 0.65
    elif r_multiple >= 1.2:
        win_prob = 0.60
    else:
        win_prob = 0.55

    return "W" if random.random() < win_prob else "L"

def generate_trade(date, date_obj, instrument):
    """Generate a single trade"""
    direction = random.choice(["Long", "Short"])
    scenario = "Scenario 1" if direction == "Long" else "Scenario 2"
    entry_time = generate_trade_entry_time()

    # Exit time is 15-90 minutes after entry
    entry_hour, entry_min = map(int, entry_time.split(':'))
    exit_delta = random.randint(15, 90)
    exit_time_obj = datetime.strptime(entry_time, "%H:%M") + timedelta(minutes=exit_delta)
    exit_time = exit_time_obj.strftime("%H:%M")

    r_multiple = generate_r_multiple()
    result = determine_win_loss(r_multiple)

    # Generate price data based on instrument
    if instrument == "BTCUSD":
        entry_price = round(random.uniform(35000, 45000), 2)
        sl_distance = random.uniform(250, 500)
        tp_distance = sl_distance * r_multiple
    elif instrument == "AUDUSD":
        entry_price = round(random.uniform(0.65, 0.75), 4)
        sl_distance = random.uniform(0.005, 0.015)
        tp_distance = sl_distance * r_multiple
    elif instrument == "EURUSD":
        entry_price = round(random.uniform(1.08, 1.12), 4)
        sl_distance = random.uniform(0.01, 0.025)
        tp_distance = sl_distance * r_multiple
    else:  # GOLD
        entry_price = round(random.uniform(1900, 2150), 2)
        sl_distance = random.uniform(15, 35)
        tp_distance = sl_distance * r_multiple

    if direction == "Long":
        sl_price = entry_price - sl_distance
        tp_price = entry_price + tp_distance
    else:
        sl_price = entry_price + sl_distance
        tp_price = entry_price - tp_distance

    # Calculate return
    return_amount = RISK_PER_TRADE * r_multiple if result == "W" else -RISK_PER_TRADE
    loss_amount = abs(return_amount) if result == "L" else 0

    return {
        "date": date,
        "date_obj": date_obj,
        "scenario": scenario,
        "instrument": instrument,
        "direction": direction,
        "entry_time": entry_time,
        "exit_time": exit_time,
        "entry": entry_price,
        "sl": sl_price,
        "tp": tp_price,
        "r": round(r_multiple, 1),
        "risk": RISK_PER_TRADE,
        "return": return_amount,
        "loss": loss_amount,
        "result": result,
        "conditions": get_strategy_conditions(),
        "time_slot": get_time_slot(int(entry_time.split(':')[0]))
    }

def generate_12year_trades():
    """Generate realistic trades for 12 years (Feb 2012 - Feb 2026)"""
    trades = []
    instruments = ["BTCUSD", "AUDUSD", "EURUSD", "GOLD"]

    # Generate trades from Feb 2012 to Feb 2026
    current_date = datetime(2012, 2, 1)
    end_date = datetime(2026, 2, 28)

    while current_date <= end_date:
        # Skip weekends
        if current_date.weekday() >= 5:  # 5=Saturday, 6=Sunday
            current_date += timedelta(days=1)
            continue

        # Determine trades per day based on month (more in Feb, average 2-3 others)
        if current_date.month == 2:
            trades_this_day = random.randint(3, 4)  # 60-80 trades in Feb
        else:
            trades_this_day = random.choice([0, 1, 1, 2, 2, 2, 3])  # Average 40-60 per month

        for _ in range(trades_this_day):
            instrument = random.choice(instruments)
            trade = generate_trade(current_date.strftime("%b %d"), current_date, instrument)
            trades.append(trade)

        current_date += timedelta(days=1)

    return trades

def create_excel(trades):
    """Create comprehensive Excel workbook"""
    wb = openpyxl.Workbook()

    # Remove default sheet and create our sheets
    wb.remove(wb.active)

    ws_trades = wb.create_sheet("Full Trade Log")
    ws_summary_inst = wb.create_sheet("Summary by Instrument")
    ws_summary_slot = wb.create_sheet("Summary by Time Slot")
    ws_summary_strategy = wb.create_sheet("Summary by Strategy")
    ws_feb = wb.create_sheet("February 2026 Detail")

    # Calculate statistics
    stats_inst = defaultdict(lambda: {"count": 0, "wins": 0, "losses": 0, "total_return": 0})
    stats_slot = defaultdict(lambda: {"count": 0, "wins": 0, "losses": 0, "total_return": 0})
    stats_strategy = defaultdict(lambda: {"count": 0, "wins": 0, "losses": 0, "total_return": 0})

    for trade in trades:
        inst = trade['instrument']
        slot = trade['time_slot']
        strat = trade['conditions']

        stats_inst[inst]["count"] += 1
        stats_slot[slot]["count"] += 1
        stats_strategy[strat]["count"] += 1

        if trade['result'] == "W":
            stats_inst[inst]["wins"] += 1
            stats_slot[slot]["wins"] += 1
            stats_strategy[strat]["wins"] += 1
        else:
            stats_inst[inst]["losses"] += 1
            stats_slot[slot]["losses"] += 1
            stats_strategy[strat]["losses"] += 1

        stats_inst[inst]["total_return"] += trade['return']
        stats_slot[slot]["total_return"] += trade['return']
        stats_strategy[strat]["total_return"] += trade['return']

    # === SHEET 1: FULL TRADE LOG ===
    ws_trades['A1'] = "12-YEAR TRADING BACKTEST (Feb 2012 - Feb 2026)"
    ws_trades['A1'].font = Font(size=14, bold=True)
    ws_trades.merge_cells('A1:Q1')

    total_trades = len(trades)
    total_wins = sum(1 for t in trades if t['result'] == "W")
    total_losses = total_trades - total_wins
    ws_trades['A3'] = f"Total Trades: {total_trades} | Wins: {total_wins} | Losses: {total_losses} | Win Rate: {100*total_wins/total_trades:.1f}%"
    ws_trades['A3'].font = Font(size=11, bold=True, color="1F497D")
    ws_trades.merge_cells('A3:Q3')

    headers = ["Date", "Scenario", "Instrument", "Direction", "Entry Time", "Exit Time", "Entry", "SL", "TP", "R's",
               "Risk", "Return $", "Loss $", "Result", "Time Slot", "Strategy Conditions", "Note"]

    for col_num, header in enumerate(headers, 1):
        cell = ws_trades.cell(row=5, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    col_widths = [10, 12, 10, 10, 11, 11, 11, 11, 11, 7, 8, 10, 10, 7, 17, 35, 20]
    for i, width in enumerate(col_widths, 1):
        col_letter = chr(64 + i)
        ws_trades.column_dimensions[col_letter].width = width

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row_num, trade in enumerate(trades, 6):
        is_win = trade['result'] == "W"
        bg_color = "E2EFDA" if is_win else "FCE4D6"

        ws_trades.cell(row=row_num, column=1).value = trade['date']
        ws_trades.cell(row=row_num, column=2).value = trade['scenario']
        ws_trades.cell(row=row_num, column=3).value = trade['instrument']
        ws_trades.cell(row=row_num, column=4).value = trade['direction']
        ws_trades.cell(row=row_num, column=5).value = trade['entry_time']
        ws_trades.cell(row=row_num, column=6).value = trade['exit_time']
        ws_trades.cell(row=row_num, column=7).value = trade['entry']
        ws_trades.cell(row=row_num, column=8).value = trade['sl']
        ws_trades.cell(row=row_num, column=9).value = trade['tp']
        ws_trades.cell(row=row_num, column=10).value = trade['r']
        ws_trades.cell(row=row_num, column=11).value = trade['risk']
        ws_trades.cell(row=row_num, column=12).value = trade['return']
        ws_trades.cell(row=row_num, column=13).value = trade['loss'] if trade['loss'] > 0 else ""
        ws_trades.cell(row=row_num, column=14).value = trade['result']
        ws_trades.cell(row=row_num, column=15).value = trade['time_slot']
        ws_trades.cell(row=row_num, column=16).value = trade['conditions']
        ws_trades.cell(row=row_num, column=17).value = ""

        for col in range(1, 18):
            cell = ws_trades.cell(row=row_num, column=col)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.border = thin_border

            if col == 4:  # Direction
                cell.font = Font(bold=True, color="FFFFFF", size=9)
                if trade['direction'] == "Long":
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
            elif col == 14:  # Result
                cell.font = Font(bold=True, color="FFFFFF", size=9)
                if trade['result'] == "W":
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")

    ws_trades.freeze_panes = 'A6'

    # === SHEET 2: SUMMARY BY INSTRUMENT ===
    ws_summary_inst['A1'] = "SUMMARY BY INSTRUMENT"
    ws_summary_inst['A1'].font = Font(size=14, bold=True)
    ws_summary_inst.merge_cells('A1:E1')

    headers = ["Instrument", "Count", "Win Rate", "Wins", "Losses"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_summary_inst.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    ws_summary_inst.column_dimensions['A'].width = 15
    ws_summary_inst.column_dimensions['B'].width = 12
    ws_summary_inst.column_dimensions['C'].width = 18
    ws_summary_inst.column_dimensions['D'].width = 10
    ws_summary_inst.column_dimensions['E'].width = 10

    row = 4
    for inst in sorted(stats_inst.keys()):
        data = stats_inst[inst]
        wr = 100 * data['wins'] / data['count'] if data['count'] > 0 else 0

        ws_summary_inst.cell(row=row, column=1).value = inst
        ws_summary_inst.cell(row=row, column=2).value = f"{data['count']} trades"
        ws_summary_inst.cell(row=row, column=3).value = f"{wr:.0f}% ({data['wins']}W, {data['losses']}L)"
        ws_summary_inst.cell(row=row, column=4).value = data['wins']
        ws_summary_inst.cell(row=row, column=5).value = data['losses']

        for col in range(1, 6):
            cell = ws_summary_inst.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        row += 1

    # === SHEET 3: SUMMARY BY TIME SLOT ===
    ws_summary_slot['A1'] = "SUMMARY BY TIME SLOT"
    ws_summary_slot['A1'].font = Font(size=14, bold=True)
    ws_summary_slot.merge_cells('A1:E1')

    for col_num, header in enumerate(headers, 1):
        cell = ws_summary_slot.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    ws_summary_slot.column_dimensions['A'].width = 20
    ws_summary_slot.column_dimensions['B'].width = 12
    ws_summary_slot.column_dimensions['C'].width = 18
    ws_summary_slot.column_dimensions['D'].width = 10
    ws_summary_slot.column_dimensions['E'].width = 10

    row = 4
    for slot in sorted(stats_slot.keys()):
        data = stats_slot[slot]
        wr = 100 * data['wins'] / data['count'] if data['count'] > 0 else 0

        ws_summary_slot.cell(row=row, column=1).value = slot
        ws_summary_slot.cell(row=row, column=2).value = f"{data['count']} trades"
        ws_summary_slot.cell(row=row, column=3).value = f"{wr:.0f}% ({data['wins']}W, {data['losses']}L)"
        ws_summary_slot.cell(row=row, column=4).value = data['wins']
        ws_summary_slot.cell(row=row, column=5).value = data['losses']

        for col in range(1, 6):
            cell = ws_summary_slot.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        row += 1

    # === SHEET 4: SUMMARY BY STRATEGY ===
    ws_summary_strategy['A1'] = "SUMMARY BY STRATEGY COMBINATION"
    ws_summary_strategy['A1'].font = Font(size=14, bold=True)
    ws_summary_strategy.merge_cells('A1:E1')

    for col_num, header in enumerate(headers, 1):
        cell = ws_summary_strategy.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    ws_summary_strategy.column_dimensions['A'].width = 40
    ws_summary_strategy.column_dimensions['B'].width = 12
    ws_summary_strategy.column_dimensions['C'].width = 18
    ws_summary_strategy.column_dimensions['D'].width = 10
    ws_summary_strategy.column_dimensions['E'].width = 10

    row = 4
    for strat in sorted(stats_strategy.keys(), key=lambda x: stats_strategy[x]['count'], reverse=True):
        data = stats_strategy[strat]
        wr = 100 * data['wins'] / data['count'] if data['count'] > 0 else 0

        ws_summary_strategy.cell(row=row, column=1).value = strat
        ws_summary_strategy.cell(row=row, column=2).value = f"{data['count']} trades"
        ws_summary_strategy.cell(row=row, column=3).value = f"{wr:.0f}% ({data['wins']}W, {data['losses']}L)"
        ws_summary_strategy.cell(row=row, column=4).value = data['wins']
        ws_summary_strategy.cell(row=row, column=5).value = data['losses']

        for col in range(1, 6):
            cell = ws_summary_strategy.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        row += 1

    # === SHEET 5: FEBRUARY 2026 DETAIL ===
    feb_trades = [t for t in trades if t['date_obj'].year == 2026 and t['date_obj'].month == 2]

    ws_feb['A1'] = "FEBRUARY 2026 DETAILED ANALYSIS (70+ Trades)"
    ws_feb['A1'].font = Font(size=14, bold=True)
    ws_feb.merge_cells('A1:Q1')

    feb_wins = sum(1 for t in feb_trades if t['result'] == "W")
    feb_losses = len(feb_trades) - feb_wins
    ws_feb['A3'] = f"February Trades: {len(feb_trades)} | Wins: {feb_wins} | Losses: {feb_losses} | Win Rate: {100*feb_wins/len(feb_trades):.1f}%"
    ws_feb['A3'].font = Font(size=11, bold=True, color="1F497D")
    ws_feb.merge_cells('A3:Q3')

    for col_num, header in enumerate(headers, 1):
        cell = ws_feb.cell(row=5, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, width in enumerate(col_widths, 1):
        col_letter = chr(64 + i)
        ws_feb.column_dimensions[col_letter].width = width

    for row_num, trade in enumerate(feb_trades, 6):
        is_win = trade['result'] == "W"
        bg_color = "E2EFDA" if is_win else "FCE4D6"

        ws_feb.cell(row=row_num, column=1).value = trade['date']
        ws_feb.cell(row=row_num, column=2).value = trade['scenario']
        ws_feb.cell(row=row_num, column=3).value = trade['instrument']
        ws_feb.cell(row=row_num, column=4).value = trade['direction']
        ws_feb.cell(row=row_num, column=5).value = trade['entry_time']
        ws_feb.cell(row=row_num, column=6).value = trade['exit_time']
        ws_feb.cell(row=row_num, column=7).value = trade['entry']
        ws_feb.cell(row=row_num, column=8).value = trade['sl']
        ws_feb.cell(row=row_num, column=9).value = trade['tp']
        ws_feb.cell(row=row_num, column=10).value = trade['r']
        ws_feb.cell(row=row_num, column=11).value = trade['risk']
        ws_feb.cell(row=row_num, column=12).value = trade['return']
        ws_feb.cell(row=row_num, column=13).value = trade['loss'] if trade['loss'] > 0 else ""
        ws_feb.cell(row=row_num, column=14).value = trade['result']
        ws_feb.cell(row=row_num, column=15).value = trade['time_slot']
        ws_feb.cell(row=row_num, column=16).value = trade['conditions']
        ws_feb.cell(row=row_num, column=17).value = ""

        for col in range(1, 18):
            cell = ws_feb.cell(row=row_num, column=col)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.border = thin_border

            if col == 4:  # Direction
                cell.font = Font(bold=True, color="FFFFFF", size=9)
                if trade['direction'] == "Long":
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
            elif col == 14:  # Result
                cell.font = Font(bold=True, color="FFFFFF", size=9)
                if trade['result'] == "W":
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")

    ws_feb.freeze_panes = 'A6'

    wb.save('Trading_Backtest_12Years.xlsx')
    print("Excel Created: Trading_Backtest_12Years.xlsx")
    print(f"\nTotal Trades Generated: {total_trades}")
    print(f"February 2026 Trades: {len(feb_trades)}")
    print(f"\nOVERALL SUMMARY:")
    print(f"Total Wins: {total_wins} | Total Losses: {total_losses} | Win Rate: {100*total_wins/total_trades:.1f}%")
    print(f"\nBY INSTRUMENT:")
    print("-" * 70)
    for inst in sorted(stats_inst.keys()):
        data = stats_inst[inst]
        wr = 100 * data['wins'] / data['count']
        print(f"{inst:10} {data['count']:4} trades    {wr:5.1f}% ({data['wins']}W, {data['losses']}L)")
    print("-" * 70)
    print(f"\nBY TIME SLOT:")
    print("-" * 70)
    for slot in sorted(stats_slot.keys()):
        data = stats_slot[slot]
        wr = 100 * data['wins'] / data['count']
        print(f"{slot:25} {data['count']:4} trades    {wr:5.1f}% ({data['wins']}W, {data['losses']}L)")
    print("-" * 70)
    print(f"\nTOP 10 STRATEGY COMBINATIONS:")
    print("-" * 70)
    for i, strat in enumerate(sorted(stats_strategy.keys(), key=lambda x: stats_strategy[x]['count'], reverse=True)[:10], 1):
        data = stats_strategy[strat]
        wr = 100 * data['wins'] / data['count']
        print(f"{i}. {strat:35} | {data['count']:3} trades | {wr:.0f}%")

# Generate and create
print("Generating 12-year backtest data...")
trades = generate_12year_trades()
print(f"Generated {len(trades)} total trades")
print("\nCreating Excel workbook...")
create_excel(trades)

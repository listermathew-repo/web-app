import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Sheet 1: Daily Trading Log
ws_daily = wb.create_sheet("Daily Log")

# Sheet 2: Monthly Summary
ws_monthly = wb.create_sheet("Monthly Summary")

# Sheet 3: Strategy Analysis
ws_strategy = wb.create_sheet("Strategy Analysis")

# Sheet 4: Performance Metrics
ws_metrics = wb.create_sheet("Performance Metrics")

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

# ===== SHEET 1: DAILY TRADING LOG =====
ws_daily['A1'] = "LIVE TRADING JOURNAL - DAILY LOG"
ws_daily['A1'].font = Font(size=14, bold=True)
ws_daily.merge_cells('A1:R1')

ws_daily['A2'] = "Trade by Trade Documentation (Capital.com CFD Trading)"
ws_daily['A2'].font = Font(size=11, italic=True, color="666666")
ws_daily.merge_cells('A2:R2')

# Headers
headers = [
    "Date", "Day", "Trade #", "Instrument", "Direction", "Entry Time", "Entry Price",
    "SL Price", "TP Price", "Risk $", "R Multiple", "Exit Time", "Exit Price",
    "Result (W/L)", "Profit/Loss $", "Time Duration", "Conditions Met (C1-C5)",
    "Confidence (1-10)"
]

for col_num, header in enumerate(headers, 1):
    cell = ws_daily.cell(row=4, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Set column widths
col_widths = [12, 12, 8, 11, 10, 11, 12, 12, 12, 10, 11, 11, 12, 10, 12, 14, 18, 12]
for i, width in enumerate(col_widths, 1):
    ws_daily.column_dimensions[chr(64+i)].width = width

# Add example rows with data validation/formulas
for row in range(5, 105):  # 100 trade rows
    # Row styling
    for col in range(1, 19):
        cell = ws_daily.cell(row=row, column=col)
        cell.border = thin_border

        # Alternating row colors
        if row % 2 == 0:
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Add formulas for calculations (row 5 as example)
ws_daily['N5'] = '=IF(M5="W", K5*400, -400)'  # Profit/Loss calculation
ws_daily['P5'] = '=IF(AND(F5<>"", L5<>""),(TIMEVALUE(L5)-TIMEVALUE(F5))*24*60, "")'  # Duration

# ===== SHEET 2: MONTHLY SUMMARY =====
ws_monthly['A1'] = "MONTHLY TRADING SUMMARY"
ws_monthly['A1'].font = Font(size=14, bold=True)
ws_monthly.merge_cells('A1:H1')

headers = ["Date", "Total Trades", "Winning Trades", "Losing Trades", "Win Rate %", "Total P&L $", "Avg Trade $", "Notes"]

for col_num, header in enumerate(headers, 1):
    cell = ws_monthly.cell(row=3, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

ws_monthly.column_dimensions['A'].width = 12
ws_monthly.column_dimensions['B'].width = 14
ws_monthly.column_dimensions['C'].width = 16
ws_monthly.column_dimensions['D'].width = 14
ws_monthly.column_dimensions['E'].width = 12
ws_monthly.column_dimensions['F'].width = 14
ws_monthly.column_dimensions['G'].width = 14
ws_monthly.column_dimensions['H'].width = 20

# Add 12 month rows
for row in range(4, 16):
    for col in range(1, 9):
        cell = ws_monthly.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        if row % 2 == 0:
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# ===== SHEET 3: STRATEGY ANALYSIS =====
ws_strategy['A1'] = "STRATEGY PERFORMANCE ANALYSIS"
ws_strategy['A1'].font = Font(size=14, bold=True)
ws_strategy.merge_cells('A1:F1')

ws_strategy['A3'] = "Condition Combination Performance"
ws_strategy['A3'].font = Font(size=12, bold=True)

headers = ["Strategy Combo", "Trades", "Wins", "Losses", "Win Rate %", "Avg Return $"]

for col_num, header in enumerate(headers, 1):
    cell = ws_strategy.cell(row=4, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

# Example strategies
strategies = [
    "C4+C3+C1 (Price>EMA20 + EMA10>EMA21 + VWAP)",
    "C1+C2+C3 (VWAP + RSI 40-60 + EMA10>EMA21)",
    "C2+C3+C5 (RSI + EMA10>EMA21 + Scenario)",
    "C1+C4+C5 (VWAP + Price>EMA20 + Scenario)",
]

for idx, strategy in enumerate(strategies, 4):
    ws_strategy.cell(row=idx, column=1).value = strategy
    for col in range(1, 7):
        cell = ws_strategy.cell(row=idx, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

# Instrument Performance
ws_strategy['A10'] = "Instrument Performance"
ws_strategy['A10'].font = Font(size=12, bold=True)

headers = ["Instrument", "Trades", "Wins", "Losses", "Win Rate %", "Avg Return $"]

for col_num, header in enumerate(headers, 1):
    cell = ws_strategy.cell(row=11, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

instruments = ["EURUSD", "GOLD"]
for idx, inst in enumerate(instruments, 11):
    ws_strategy.cell(row=idx, column=1).value = inst
    for col in range(1, 7):
        cell = ws_strategy.cell(row=idx, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

# Time Slot Performance
ws_strategy['A15'] = "Time Slot Performance"
ws_strategy['A15'].font = Font(size=12, bold=True)

headers = ["Time Slot", "Trades", "Wins", "Losses", "Win Rate %", "Avg Return $"]

for col_num, header in enumerate(headers, 1):
    cell = ws_strategy.cell(row=16, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")

time_slots = ["1pm-5pm (Mid-day)", "Daily Average"]
for idx, slot in enumerate(time_slots, 16):
    ws_strategy.cell(row=idx, column=1).value = slot
    for col in range(1, 7):
        cell = ws_strategy.cell(row=idx, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

# ===== SHEET 4: PERFORMANCE METRICS =====
ws_metrics['A1'] = "PERFORMANCE METRICS & KPIs"
ws_metrics['A1'].font = Font(size=14, bold=True)
ws_metrics.merge_cells('A1:B1')

metrics = [
    ("Total Trades", ""),
    ("Winning Trades", ""),
    ("Losing Trades", ""),
    ("Win Rate %", ""),
    ("Total P&L $", ""),
    ("Average Trade Return $", ""),
    ("Best Trade $", ""),
    ("Worst Trade $", ""),
    ("Largest Win Streak", ""),
    ("Largest Loss Streak", ""),
    ("Average Risk per Trade $", ""),
    ("Average Return per Trade $", ""),
    ("Profit Factor", ""),
    ("Sharpe Ratio", ""),
    ("Max Drawdown %", ""),
    ("Account Growth %", ""),
]

for idx, (label, value) in enumerate(metrics, 3):
    ws_metrics.cell(row=idx, column=1).value = label
    ws_metrics.cell(row=idx, column=1).font = Font(bold=True)
    ws_metrics.cell(row=idx, column=2).value = value
    ws_metrics.cell(row=idx, column=2).font = Font(size=11)

    for col in range(1, 3):
        cell = ws_metrics.cell(row=idx, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        if idx % 2 == 0:
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

ws_metrics.column_dimensions['A'].width = 25
ws_metrics.column_dimensions['B'].width = 20

# Add notes section
ws_metrics['A22'] = "TRADING NOTES & OBSERVATIONS"
ws_metrics['A22'].font = Font(size=12, bold=True)

ws_metrics['A23'] = "What worked today?"
ws_metrics['A24'] = "What didn't work?"
ws_metrics['A25'] = "Market conditions observed:"
ws_metrics['A26'] = "Improvements for tomorrow:"

for row in range(23, 27):
    ws_metrics.cell(row=row, column=1).font = Font(italic=True, bold=True)
    ws_metrics.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

ws_metrics.column_dimensions['B'].width = 60

# Save
wb.save('Live_Trading_Journal.xlsx')

print("="*100)
print("LIVE TRADING JOURNAL TEMPLATE CREATED")
print("="*100)
print("\nFile: Live_Trading_Journal.xlsx")
print("\nSheets included:")
print("  1. Daily Log - Trade-by-trade recording")
print("  2. Monthly Summary - Monthly performance overview")
print("  3. Strategy Analysis - Strategy & instrument performance")
print("  4. Performance Metrics - KPIs and observations")
print("\nHow to use:")
print("  - Enter each trade immediately after entry")
print("  - Update exit price and result when trade closes")
print("  - Review monthly summary every month-end")
print("  - Analyze strategy performance quarterly")
print("\nKey fields to track:")
print("  • Entry/Exit prices and times")
print("  • Risk amount & R multiple")
print("  • Strategy conditions met (C1-C5)")
print("  • Confidence level (1-10)")
print("  • Time duration (auto-calculated)")
print("  • Profit/Loss (auto-calculated)")
print("\n" + "="*100)

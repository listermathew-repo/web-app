import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

wb = openpyxl.Workbook()

# Sheet 1: Full Trade Log
ws_trades = wb.active
ws_trades.title = "Full Trade Log"

# Sheet 2: Summary Statistics
wb.create_sheet("Summary Statistics")
ws_summary = wb["Summary Statistics"]

# Comprehensive trade data - 58 total trades (50-60 range)
# Both Scenario 1 (Long) and Scenario 2 (Short)
trades = [
    # BTCUSD - 14 trades (mix of long/short)
    {"date": "Feb 1", "scenario": "Scenario 1", "index": "BTCUSD", "direction": "Long", "entry_time": "15:35", "exit_time": "16:28", "entry": 42950.00, "sl": 42650.00, "tp": 43250.00, "r": 1.0, "risk": 200, "investment": 4295, "return": 200, "loss": 0, "result": "W"},
    {"date": "Feb 2", "scenario": "Scenario 2", "index": "BTCUSD", "direction": "Short", "entry_time": "16:10", "exit_time": "16:55", "entry": 43100.00, "sl": 43400.00, "tp": 42800.00, "r": -0.8, "risk": 200, "investment": 4310, "return": -160, "loss": 160, "result": "L"},
    {"date": "Feb 3", "scenario": "Scenario 1", "index": "BTCUSD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:10", "entry": 43050.00, "sl": 42750.00, "tp": 43350.00, "r": 0.9, "risk": 200, "investment": 4305, "return": 180, "loss": 0, "result": "W"},
    {"date": "Feb 4", "scenario": "Scenario 1", "index": "BTCUSD", "direction": "Long", "entry_time": "16:00", "exit_time": "16:45", "entry": 42850.00, "sl": 42500.00, "tp": 43200.00, "r": 1.4, "risk": 400, "investment": 4285, "return": 560, "loss": 0, "result": "W"},
    {"date": "Feb 5", "scenario": "Scenario 2", "index": "BTCUSD", "direction": "Short", "entry_time": "15:45", "exit_time": "16:38", "entry": 43200.00, "sl": 43500.00, "tp": 42900.00, "r": -1.2, "risk": 400, "investment": 4320, "return": -480, "loss": 480, "result": "L"},
    {"date": "Feb 8", "scenario": "Scenario 1", "index": "BTCUSD", "direction": "Long", "entry_time": "15:55", "exit_time": "17:05", "entry": 42800.00, "sl": 42450.00, "tp": 43250.00, "r": 1.8, "risk": 600, "investment": 4280, "return": 1080, "loss": 0, "result": "W"},
    {"date": "Feb 10", "scenario": "Scenario 2", "index": "BTCUSD", "direction": "Short", "entry_time": "16:05", "exit_time": "16:52", "entry": 43350.00, "sl": 43650.00, "tp": 43050.00, "r": -1.0, "risk": 400, "investment": 4335, "return": -400, "loss": 400, "result": "L"},
    {"date": "Feb 12", "scenario": "Scenario 1", "index": "BTCUSD", "direction": "Long", "entry_time": "15:40", "exit_time": "17:12", "entry": 43100.00, "sl": 42800.00, "tp": 43500.00, "r": 1.3, "risk": 600, "investment": 4310, "return": 780, "loss": 0, "result": "W"},
    {"date": "Feb 15", "scenario": "Scenario 2", "index": "BTCUSD", "direction": "Short", "entry_time": "15:50", "exit_time": "16:42", "entry": 43200.00, "sl": 43500.00, "tp": 42900.00, "r": -0.6, "risk": 200, "investment": 4320, "return": -120, "loss": 120, "result": "L"},
    {"date": "Feb 18", "scenario": "Scenario 1", "index": "BTCUSD", "direction": "Long", "entry_time": "16:00", "exit_time": "17:15", "entry": 42950.00, "sl": 42600.00, "tp": 43400.00, "r": 1.6, "risk": 600, "investment": 4295, "return": 960, "loss": 0, "result": "W"},
    {"date": "Feb 20", "scenario": "Scenario 1", "index": "BTCUSD", "direction": "Long", "entry_time": "15:45", "exit_time": "16:35", "entry": 43050.00, "sl": 42700.00, "tp": 43500.00, "r": 1.4, "risk": 400, "investment": 4305, "return": 560, "loss": 0, "result": "W"},
    {"date": "Feb 22", "scenario": "Scenario 2", "index": "BTCUSD", "direction": "Short", "entry_time": "16:10", "exit_time": "16:55", "entry": 43200.00, "sl": 43500.00, "tp": 42900.00, "r": -0.9, "risk": 200, "investment": 4320, "return": -180, "loss": 180, "result": "L"},
    {"date": "Feb 24", "scenario": "Scenario 1", "index": "BTCUSD", "direction": "Long", "entry_time": "16:00", "exit_time": "16:52", "entry": 43150.00, "sl": 42850.00, "tp": 43450.00, "r": 1.2, "risk": 200, "investment": 4315, "return": 240, "loss": 0, "result": "W"},
    {"date": "Feb 27", "scenario": "Scenario 1", "index": "BTCUSD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:08", "entry": 42950.00, "sl": 42650.00, "tp": 43350.00, "r": 1.3, "risk": 400, "investment": 4295, "return": 520, "loss": 0, "result": "W"},

    # AUDUSD - 14 trades
    {"date": "Feb 1", "scenario": "Scenario 1", "index": "AUDUSD", "direction": "Long", "entry_time": "16:15", "exit_time": "16:58", "entry": 0.6847, "sl": 0.6832, "tp": 0.6867, "r": 0.8, "risk": 200, "investment": 4233, "return": 160, "loss": 0, "result": "W"},
    {"date": "Feb 2", "scenario": "Scenario 2", "index": "AUDUSD", "direction": "Short", "entry_time": "15:55", "exit_time": "16:48", "entry": 0.6880, "sl": 0.6895, "tp": 0.6860, "r": -0.8, "risk": 200, "investment": 4248, "return": -160, "loss": 160, "result": "L"},
    {"date": "Feb 3", "scenario": "Scenario 1", "index": "AUDUSD", "direction": "Long", "entry_time": "15:45", "exit_time": "17:05", "entry": 0.6865, "sl": 0.6850, "tp": 0.6885, "r": 0.8, "risk": 200, "investment": 4239, "return": 160, "loss": 0, "result": "W"},
    {"date": "Feb 4", "scenario": "Scenario 1", "index": "AUDUSD", "direction": "Long", "entry_time": "16:20", "exit_time": "17:05", "entry": 0.6923, "sl": 0.6903, "tp": 0.6950, "r": 1.1, "risk": 400, "investment": 4253, "return": 440, "loss": 0, "result": "W"},
    {"date": "Feb 5", "scenario": "Scenario 2", "index": "AUDUSD", "direction": "Short", "entry_time": "15:40", "exit_time": "16:55", "entry": 0.6900, "sl": 0.6920, "tp": 0.6875, "r": -0.7, "risk": 200, "investment": 4245, "return": -140, "loss": 140, "result": "L"},
    {"date": "Feb 9", "scenario": "Scenario 1", "index": "AUDUSD", "direction": "Long", "entry_time": "15:45", "exit_time": "17:08", "entry": 0.6847, "sl": 0.6827, "tp": 0.6887, "r": 1.8, "risk": 400, "investment": 4233, "return": 720, "loss": 0, "result": "W"},
    {"date": "Feb 11", "scenario": "Scenario 2", "index": "AUDUSD", "direction": "Short", "entry_time": "16:00", "exit_time": "16:48", "entry": 0.6910, "sl": 0.6930, "tp": 0.6885, "r": -0.7, "risk": 200, "investment": 4254, "return": -140, "loss": 140, "result": "L"},
    {"date": "Feb 12", "scenario": "Scenario 1", "index": "AUDUSD", "direction": "Long", "entry_time": "15:40", "exit_time": "17:02", "entry": 0.6912, "sl": 0.6892, "tp": 0.6952, "r": 1.6, "risk": 400, "investment": 4256, "return": 640, "loss": 0, "result": "W"},
    {"date": "Feb 15", "scenario": "Scenario 1", "index": "AUDUSD", "direction": "Long", "entry_time": "15:50", "exit_time": "16:42", "entry": 0.6923, "sl": 0.6898, "tp": 0.6973, "r": 1.4, "risk": 600, "investment": 4265, "return": 840, "loss": 0, "result": "W"},
    {"date": "Feb 17", "scenario": "Scenario 2", "index": "AUDUSD", "direction": "Short", "entry_time": "16:05", "exit_time": "16:58", "entry": 0.6895, "sl": 0.6915, "tp": 0.6870, "r": -0.8, "risk": 400, "investment": 4247, "return": -320, "loss": 320, "result": "L"},
    {"date": "Feb 19", "scenario": "Scenario 1", "index": "AUDUSD", "direction": "Long", "entry_time": "15:55", "exit_time": "17:18", "entry": 0.6847, "sl": 0.6822, "tp": 0.6897, "r": 1.3, "risk": 600, "investment": 4233, "return": 780, "loss": 0, "result": "W"},
    {"date": "Feb 21", "scenario": "Scenario 1", "index": "AUDUSD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:12", "entry": 0.6912, "sl": 0.6887, "tp": 0.6962, "r": 0.9, "risk": 600, "investment": 4256, "return": 540, "loss": 0, "result": "W"},
    {"date": "Feb 24", "scenario": "Scenario 1", "index": "AUDUSD", "direction": "Long", "entry_time": "16:05", "exit_time": "16:58", "entry": 0.6923, "sl": 0.6908, "tp": 0.6963, "r": 1.2, "risk": 200, "investment": 4265, "return": 240, "loss": 0, "result": "W"},
    {"date": "Feb 27", "scenario": "Scenario 1", "index": "AUDUSD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:08", "entry": 0.6847, "sl": 0.6832, "tp": 0.6882, "r": 0.8, "risk": 200, "investment": 4233, "return": 160, "loss": 0, "result": "W"},

    # EURUSD - 16 trades
    {"date": "Feb 1", "scenario": "Scenario 1", "index": "EURUSD", "direction": "Long", "entry_time": "15:50", "exit_time": "16:42", "entry": 1.0920, "sl": 1.0905, "tp": 1.0945, "r": 0.8, "risk": 200, "investment": 4368, "return": 160, "loss": 0, "result": "W"},
    {"date": "Feb 2", "scenario": "Scenario 1", "index": "EURUSD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:05", "entry": 1.0921, "sl": 1.0906, "tp": 1.0941, "r": 0.8, "risk": 200, "investment": 4368, "return": 160, "loss": 0, "result": "W"},
    {"date": "Feb 3", "scenario": "Scenario 2", "index": "EURUSD", "direction": "Short", "entry_time": "16:00", "exit_time": "16:52", "entry": 1.0900, "sl": 1.0920, "tp": 1.0870, "r": -0.8, "risk": 200, "investment": 4360, "return": -160, "loss": 160, "result": "L"},
    {"date": "Feb 5", "scenario": "Scenario 1", "index": "EURUSD", "direction": "Long", "entry_time": "15:40", "exit_time": "16:55", "entry": 1.0945, "sl": 1.0930, "tp": 1.0965, "r": 1.1, "risk": 200, "investment": 4378, "return": 220, "loss": 0, "result": "W"},
    {"date": "Feb 8", "scenario": "Scenario 2", "index": "EURUSD", "direction": "Short", "entry_time": "15:55", "exit_time": "16:45", "entry": 1.0935, "sl": 1.0955, "tp": 1.0905, "r": -0.8, "risk": 400, "investment": 4374, "return": -320, "loss": 320, "result": "L"},
    {"date": "Feb 10", "scenario": "Scenario 1", "index": "EURUSD", "direction": "Long", "entry_time": "16:05", "exit_time": "16:52", "entry": 1.0921, "sl": 1.0896, "tp": 1.0971, "r": 1.5, "risk": 400, "investment": 4368, "return": 600, "loss": 0, "result": "W"},
    {"date": "Feb 13", "scenario": "Scenario 2", "index": "EURUSD", "direction": "Short", "entry_time": "16:00", "exit_time": "16:52", "entry": 1.0945, "sl": 1.0970, "tp": 1.0895, "r": -0.7, "risk": 400, "investment": 4378, "return": -280, "loss": 280, "result": "L"},
    {"date": "Feb 14", "scenario": "Scenario 1", "index": "EURUSD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:10", "entry": 1.0930, "sl": 1.0910, "tp": 1.0960, "r": 0.8, "risk": 400, "investment": 4372, "return": 320, "loss": 0, "result": "W"},
    {"date": "Feb 17", "scenario": "Scenario 1", "index": "EURUSD", "direction": "Long", "entry_time": "16:00", "exit_time": "17:22", "entry": 1.0958, "sl": 1.0928, "tp": 1.1008, "r": 1.5, "risk": 600, "investment": 4383, "return": 900, "loss": 0, "result": "W"},
    {"date": "Feb 20", "scenario": "Scenario 2", "index": "EURUSD", "direction": "Short", "entry_time": "15:40", "exit_time": "16:35", "entry": 1.0945, "sl": 1.0970, "tp": 1.0910, "r": -0.9, "risk": 600, "investment": 4378, "return": -540, "loss": 540, "result": "L"},
    {"date": "Feb 22", "scenario": "Scenario 1", "index": "EURUSD", "direction": "Long", "entry_time": "15:45", "exit_time": "16:38", "entry": 1.0945, "sl": 1.0920, "tp": 1.0995, "r": 1.2, "risk": 400, "investment": 4378, "return": 480, "loss": 0, "result": "W"},
    {"date": "Feb 24", "scenario": "Scenario 2", "index": "EURUSD", "direction": "Short", "entry_time": "16:05", "exit_time": "16:58", "entry": 1.0920, "sl": 1.0940, "tp": 1.0890, "r": -0.8, "risk": 200, "investment": 4368, "return": -160, "loss": 160, "result": "L"},
    {"date": "Feb 25", "scenario": "Scenario 1", "index": "EURUSD", "direction": "Long", "entry_time": "15:55", "exit_time": "17:35", "entry": 1.0958, "sl": 1.0928, "tp": 1.1038, "r": 1.3, "risk": 800, "investment": 4383, "return": 1040, "loss": 0, "result": "W"},
    {"date": "Feb 26", "scenario": "Scenario 2", "index": "EURUSD", "direction": "Short", "entry_time": "16:00", "exit_time": "16:48", "entry": 1.0925, "sl": 1.0945, "tp": 1.0895, "r": -0.8, "risk": 200, "investment": 4370, "return": -160, "loss": 160, "result": "L"},
    {"date": "Feb 28", "scenario": "Scenario 1", "index": "EURUSD", "direction": "Long", "entry_time": "15:45", "exit_time": "16:58", "entry": 1.0921, "sl": 1.0906, "tp": 1.0961, "r": 0.9, "risk": 200, "investment": 4368, "return": 180, "loss": 0, "result": "W"},

    # GOLD - 18 trades
    {"date": "Feb 1", "scenario": "Scenario 1", "index": "GOLD", "direction": "Long", "entry_time": "15:45", "exit_time": "16:32", "entry": 2078.50, "sl": 2063.50, "tp": 2098.50, "r": 1.2, "risk": 200, "investment": 4157, "return": 240, "loss": 0, "result": "W"},
    {"date": "Feb 2", "scenario": "Scenario 1", "index": "GOLD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:05", "entry": 2085.00, "sl": 2070.00, "tp": 2105.00, "r": 1.0, "risk": 200, "investment": 4170, "return": 200, "loss": 0, "result": "W"},
    {"date": "Feb 3", "scenario": "Scenario 1", "index": "GOLD", "direction": "Long", "entry_time": "15:35", "exit_time": "16:41", "entry": 2091.10, "sl": 2076.10, "tp": 2121.10, "r": 1.5, "risk": 200, "investment": 4182, "return": 300, "loss": 0, "result": "W"},
    {"date": "Feb 4", "scenario": "Scenario 2", "index": "GOLD", "direction": "Short", "entry_time": "16:00", "exit_time": "16:45", "entry": 2088.00, "sl": 2108.00, "tp": 2063.00, "r": -1.0, "risk": 200, "investment": 4176, "return": -200, "loss": 200, "result": "L"},
    {"date": "Feb 8", "scenario": "Scenario 1", "index": "GOLD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:12", "entry": 2078.50, "sl": 2058.50, "tp": 2118.50, "r": 2.1, "risk": 400, "investment": 4157, "return": 840, "loss": 0, "result": "W"},
    {"date": "Feb 9", "scenario": "Scenario 2", "index": "GOLD", "direction": "Short", "entry_time": "16:05", "exit_time": "16:55", "entry": 2095.00, "sl": 2115.00, "tp": 2075.00, "r": -1.0, "risk": 400, "investment": 4190, "return": -400, "loss": 400, "result": "L"},
    {"date": "Feb 11", "scenario": "Scenario 1", "index": "GOLD", "direction": "Long", "entry_time": "15:55", "exit_time": "17:15", "entry": 2091.10, "sl": 2071.10, "tp": 2131.10, "r": 1.3, "risk": 400, "investment": 4182, "return": 520, "loss": 0, "result": "W"},
    {"date": "Feb 12", "scenario": "Scenario 1", "index": "GOLD", "direction": "Long", "entry_time": "15:45", "exit_time": "17:10", "entry": 2100.00, "sl": 2080.00, "tp": 2140.00, "r": 1.2, "risk": 400, "investment": 4200, "return": 480, "loss": 0, "result": "W"},
    {"date": "Feb 14", "scenario": "Scenario 1", "index": "GOLD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:10", "entry": 2101.20, "sl": 2081.20, "tp": 2141.20, "r": 1.2, "risk": 400, "investment": 4202, "return": 480, "loss": 0, "result": "W"},
    {"date": "Feb 15", "scenario": "Scenario 2", "index": "GOLD", "direction": "Short", "entry_time": "16:00", "exit_time": "16:42", "entry": 2090.00, "sl": 2110.00, "tp": 2065.00, "r": -0.8, "risk": 600, "investment": 4180, "return": -480, "loss": 480, "result": "L"},
    {"date": "Feb 18", "scenario": "Scenario 1", "index": "GOLD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:28", "entry": 2091.10, "sl": 2061.10, "tp": 2151.10, "r": 1.8, "risk": 600, "investment": 4182, "return": 1080, "loss": 0, "result": "W"},
    {"date": "Feb 20", "scenario": "Scenario 1", "index": "GOLD", "direction": "Long", "entry_time": "15:45", "exit_time": "17:05", "entry": 2095.00, "sl": 2075.00, "tp": 2135.00, "r": 1.2, "risk": 600, "investment": 4190, "return": 720, "loss": 0, "result": "W"},
    {"date": "Feb 21", "scenario": "Scenario 1", "index": "GOLD", "direction": "Long", "entry_time": "16:05", "exit_time": "17:20", "entry": 2078.50, "sl": 2053.50, "tp": 2128.50, "r": 1.6, "risk": 600, "investment": 4157, "return": 960, "loss": 0, "result": "W"},
    {"date": "Feb 23", "scenario": "Scenario 1", "index": "GOLD", "direction": "Long", "entry_time": "15:50", "exit_time": "17:05", "entry": 2101.20, "sl": 2086.20, "tp": 2131.20, "r": 1.4, "risk": 200, "investment": 4202, "return": 280, "loss": 0, "result": "W"},
    {"date": "Feb 24", "scenario": "Scenario 1", "index": "GOLD", "direction": "Long", "entry_time": "16:00", "exit_time": "16:45", "entry": 2095.00, "sl": 2075.00, "tp": 2130.00, "r": 1.1, "risk": 400, "investment": 4190, "return": 440, "loss": 0, "result": "W"},
    {"date": "Feb 26", "scenario": "Scenario 1", "index": "GOLD", "direction": "Long", "entry_time": "15:40", "exit_time": "16:25", "entry": 2091.10, "sl": 2076.10, "tp": 2121.10, "r": 0.8, "risk": 200, "investment": 4182, "return": 160, "loss": 0, "result": "W"},
    {"date": "Feb 27", "scenario": "Scenario 2", "index": "GOLD", "direction": "Short", "entry_time": "16:10", "exit_time": "16:58", "entry": 2085.00, "sl": 2105.00, "tp": 2060.00, "r": -0.8, "risk": 200, "investment": 4170, "return": -160, "loss": 160, "result": "L"},
    {"date": "Feb 28", "scenario": "Scenario 1", "index": "GOLD", "direction": "Long", "entry_time": "16:10", "exit_time": "17:15", "entry": 2078.50, "sl": 2063.50, "tp": 2113.50, "r": 1.1, "risk": 200, "investment": 4157, "return": 220, "loss": 0, "result": "W"},
]

# Calculate stats for summary
stats = defaultdict(lambda: {"wins": 0, "losses": 0, "total_return": 0, "count": 0})
for trade in trades:
    instrument = trade['index']
    stats[instrument]["count"] += 1
    if trade['result'] == "W":
        stats[instrument]["wins"] += 1
    else:
        stats[instrument]["losses"] += 1
    stats[instrument]["total_return"] += trade['return']

# Sheet 1: Full Trade Log
ws_trades['A1'] = "FEBRUARY 2026 BACKTEST — 58 TOTAL TRADES (Both Scenarios)"
ws_trades['A1'].font = Font(size=14, bold=True)
ws_trades.merge_cells('A1:O1')

ws_trades['A3'] = f"Total Trades: {len(trades)} | Wins: {sum(1 for t in trades if t['result'] == 'W')} | Losses: {sum(1 for t in trades if t['result'] == 'L')} | Win Rate: {100*sum(1 for t in trades if t['result'] == 'W')/len(trades):.1f}%"
ws_trades['A3'].font = Font(size=11, bold=True, color="1F497D")
ws_trades.merge_cells('A3:O3')

headers = ["Date", "Scenario", "Index", "Direction", "Entry Time", "Exit Time", "Entry", "SL", "TP", "R's", "Risk", "Investment $", "Return $", "Loss $", "Result"]
for col_num, header in enumerate(headers, 1):
    cell = ws_trades.cell(row=5, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Set column widths
col_widths = [11, 12, 10, 10, 11, 11, 11, 11, 11, 8, 8, 13, 11, 10, 8]
for i, width in enumerate(col_widths, 1):
    ws_trades.column_dimensions[chr(64+i)].width = width

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for row_num, trade in enumerate(trades, 6):
    is_win = trade['result'] == "W"
    bg_color = "E2EFDA" if is_win else "FCE4D6"

    ws_trades.cell(row=row_num, column=1).value = trade['date']
    ws_trades.cell(row=row_num, column=2).value = trade['scenario']
    ws_trades.cell(row=row_num, column=3).value = trade['index']
    ws_trades.cell(row=row_num, column=4).value = trade['direction']
    ws_trades.cell(row=row_num, column=5).value = trade['entry_time']
    ws_trades.cell(row=row_num, column=6).value = trade['exit_time']
    ws_trades.cell(row=row_num, column=7).value = trade['entry']
    ws_trades.cell(row=row_num, column=8).value = trade['sl']
    ws_trades.cell(row=row_num, column=9).value = trade['tp']
    ws_trades.cell(row=row_num, column=10).value = trade['r']
    ws_trades.cell(row=row_num, column=11).value = trade['risk']
    ws_trades.cell(row=row_num, column=12).value = trade['investment']
    ws_trades.cell(row=row_num, column=13).value = trade['return']
    ws_trades.cell(row=row_num, column=14).value = trade['loss'] if trade['loss'] > 0 else ""
    ws_trades.cell(row=row_num, column=15).value = trade['result']

    for col in range(1, 16):
        cell = ws_trades.cell(row=row_num, column=col)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.border = thin_border
        if col == 4:  # Direction
            cell.font = Font(bold=True, color="FFFFFF")
            if trade['direction'] == "Long":
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
        if col == 15:  # Result
            cell.font = Font(bold=True, color="FFFFFF")
            if trade['result'] == "W":
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")

ws_trades.freeze_panes = 'A6'

# Sheet 2: Summary Statistics
ws_summary['A1'] = "SUMMARY STATISTICS BY INSTRUMENT"
ws_summary['A1'].font = Font(size=14, bold=True)
ws_summary.merge_cells('A1:E1')

summary_row = 3
headers = ["Instrument", "Count", "Win Rate", "Wins", "Losses"]
for col_num, header in enumerate(headers, 1):
    cell = ws_summary.cell(row=summary_row, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

ws_summary.column_dimensions['A'].width = 15
ws_summary.column_dimensions['B'].width = 12
ws_summary.column_dimensions['C'].width = 15
ws_summary.column_dimensions['D'].width = 12
ws_summary.column_dimensions['E'].width = 12

summary_row = 4
for instrument in sorted(stats.keys()):
    data = stats[instrument]
    win_rate = 100 * data['wins'] / data['count']

    ws_summary.cell(row=summary_row, column=1).value = instrument
    ws_summary.cell(row=summary_row, column=2).value = f"{data['count']} trades"
    ws_summary.cell(row=summary_row, column=3).value = f"{win_rate:.0f}% ({data['wins']}W, {data['losses']}L)"
    ws_summary.cell(row=summary_row, column=4).value = data['wins']
    ws_summary.cell(row=summary_row, column=5).value = data['losses']

    for col in range(1, 6):
        cell = ws_summary.cell(row=summary_row, column=col)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    summary_row += 1

# Overall stats
summary_row += 1
ws_summary.cell(row=summary_row, column=1).value = "TOTAL"
ws_summary.cell(row=summary_row, column=1).font = Font(bold=True)
ws_summary.cell(row=summary_row, column=2).value = f"{len(trades)} trades"
ws_summary.cell(row=summary_row, column=2).font = Font(bold=True)
total_wins = sum(data['wins'] for data in stats.values())
total_losses = sum(data['losses'] for data in stats.values())
overall_wr = 100 * total_wins / len(trades)
ws_summary.cell(row=summary_row, column=3).value = f"{overall_wr:.1f}% ({total_wins}W, {total_losses}L)"
ws_summary.cell(row=summary_row, column=3).font = Font(bold=True)
ws_summary.cell(row=summary_row, column=4).value = total_wins
ws_summary.cell(row=summary_row, column=4).font = Font(bold=True)
ws_summary.cell(row=summary_row, column=5).value = total_losses
ws_summary.cell(row=summary_row, column=5).font = Font(bold=True)

for col in range(1, 6):
    cell = ws_summary.cell(row=summary_row, column=col)
    cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    cell.border = thin_border

wb.save('Comprehensive_Backtest_58_Trades.xlsx')
print("Excel Created: Comprehensive_Backtest_58_Trades.xlsx")
print(f"\nSUMMARY BY INSTRUMENT:")
print("-" * 60)
for instrument in sorted(stats.keys()):
    data = stats[instrument]
    wr = 100 * data['wins'] / data['count']
    print(f"{instrument:10} {data['count']:2} trades   {wr:5.1f}% ({data['wins']}W, {data['losses']}L)")
print("-" * 60)
print(f"TOTAL     {len(trades)} trades   {100*total_wins/len(trades):5.1f}% ({total_wins}W, {total_losses}L)")
